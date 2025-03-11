import pandas as pd
import tkinter as tk
from tkinter import messagebox, simpledialog, filedialog
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from document_generator import create_leave_application


class LeaveApplicationGUI:
    def __init__(self, root):
        self.root = root
        self.root.geometry("450x450")
        self.root.title("假条生成器")
        self.all_info = []
        self.current_person = 0
        self.num_people = 0
        self.default_reason = ""
        self.default_date = {}

    def show_format_hint(self):
        hint_message = (
            "请确保表格包含以下字段：\n"
            "- 学院\n"
            "- 专业\n"
            "- 班级\n"
            "- 姓名\n"
            "- 学号\n"
            "- 事由\n"
            "- 请假时间\n"
            "- 日期（年）\n"
            "- 日期（月）\n"
            "- 日期（日）"
        )
        messagebox.showinfo("表格格式要求", hint_message)

    def generate_empty_table(self):
        columns = ['学院', '专业', '班级', '姓名', '学号', '事由', '请假时间', '日期（年）', '日期（月）', '日期（日）']
        df = pd.DataFrame(columns=columns)
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel文件", "*.xlsx")])
        if not file_path:
            return

        try:
            df.to_excel(file_path, index=False)
            workbook = load_workbook(file_path)
            sheet = workbook.active

            common_alignment = Alignment(horizontal='center', vertical='center')
            common_font = Font(name='仿宋_GB2312')
            header_font = Font(name='仿宋_GB2312', bold=True, size=14)

            # 设置标题行样式
            for cell in sheet[1]:
                cell.font = header_font
                cell.alignment = common_alignment

            # 设置数据行样式
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.font = common_font
                    cell.alignment = common_alignment

            # 调整部分列宽
            column_widths = {'A': 20, 'B': 15, 'C': 20, 'F': 30, 'G': 30, 'H': 15, 'I': 15, 'J': 15}
            for col, width in column_widths.items():
                sheet.column_dimensions[col].width = width

            workbook.save(file_path)
            messagebox.showinfo("成功", f"空表格已生成并保存到: {file_path}")
        except Exception as e:
            messagebox.showerror("错误", f"无法生成空表格: {e}")

    def table_input(self):
        self.show_format_hint()
        file_path = filedialog.askopenfilename(filetypes=[("Excel文件", "*.xlsx;*.xls")])
        if not file_path:
            return
        try:
            df = pd.read_excel(file_path)
            required_columns = ['学院', '专业', '班级', '姓名', '学号', '事由', '请假时间', '日期（年）', '日期（月）', '日期（日）']
            if not all(col in df.columns for col in required_columns):
                messagebox.showerror("错误", "Excel文件缺少必要的列,请确保包含所有所需的字段。")
                return

            # 转换所有字段为字符串以防止 replace() 出现错误
            for col in df.columns:
                df[col] = df[col].astype(str)

            self.all_info = []
            incomplete_info = []
            for _, row in df.iterrows():
                info = row.to_dict()
                if all(info.get(col, '').strip() for col in required_columns):
                    self.all_info.append(info)
                else:
                    incomplete_info.append(info)

            if incomplete_info:
                messagebox.showwarning("信息不完整", f"发现 {len(incomplete_info)} 条信息不完整，这些信息将不会被填入。")
                for idx, incomplete in enumerate(incomplete_info, start=1):
                    print(f"不完整信息 {idx}: {incomplete}")

            if self.all_info:
                messagebox.showinfo("成功", f"从表格中导入了 {len(self.all_info)} 条完整的假条信息。")
                create_leave_application(self.all_info)
            else:
                messagebox.showinfo("提示", "没有完整的信息被导入。")
        except Exception as e:
            messagebox.showerror("错误", f"无法读取文件: {e}")

    def show_text_input_page(self):
        self.num_people = simpledialog.askinteger("人数输入", "请输入假条人数：")
        if not self.num_people or self.num_people <= 0:
            messagebox.showerror("输入错误", "人数必须大于0!")
            return

        self.default_reason = simpledialog.askstring("批量设置", "请输入假条的事由（留空则跳过）：")
        year = simpledialog.askinteger("批量设置", "请输入假条的日期（年）：", initialvalue=2024)
        month = simpledialog.askinteger("批量设置", "请输入假条的日期（月）：", initialvalue=1)
        day = simpledialog.askinteger("批量设置", "请输入假条的日期（日）：", initialvalue=1)
        if year and month and day:
            self.default_date = {"日期（年）": str(year), "日期（月）": str(month), "日期（日）": str(day)}
        else:
            self.default_date = {}

        self.current_person = 0
        self.all_info = []
        self.display_input_fields()

    def display_input_fields(self):
        for widget in self.root.winfo_children():
            widget.destroy()

        fields = ['学院', '专业', '班级', '姓名', '学号', '事由', '请假时间', '日期（年）', '日期（月）', '日期（日）']
        entries = {}

        tk.Label(self.root, text=f"请输入第 {self.current_person + 1} 个假条的信息：", font=("仿宋GB_2312", 14)).grid(row=0, columnspan=2, pady=10)

        for j, field in enumerate(fields):
            tk.Label(self.root, text=field + ":", font=("仿宋GB_2312", 12)).grid(row=j + 1, column=0, padx=5, pady=2, sticky='e')
            entry = tk.Entry(self.root, width=30, font=("仿宋GB_2312", 12))
            entry.grid(row=j + 1, column=1, padx=5, pady=2)
            if field == '事由' and self.default_reason:
                entry.insert(0, self.default_reason)
            elif field in self.default_date:
                entry.insert(0, self.default_date[field])
            entries[field] = entry

        submit_button = tk.Button(
            self.root, text="保存并继续",
            command=lambda: self.submit(entries, fields),
            width=15, height=2, font=("仿宋GB_2312", 12)
        )
        submit_button.grid(row=len(fields) + 1, columnspan=2, pady=20)

    def submit(self, entries, fields):
        info = {field: entries[field].get().strip() for field in fields}
        self.all_info.append(info)
        self.current_person += 1

        if self.current_person < self.num_people:
            self.display_input_fields()
        else:
            messagebox.showinfo("成功", "所有信息已保存，生成假条！")
            create_leave_application(self.all_info)

    def show_main_page(self):
        for widget in self.root.winfo_children():
            widget.destroy()
        button_frame = tk.Frame(self.root)
        button_frame.pack(expand=True)
        tk.Button(
            button_frame, text="表格输入",
            command=self.table_input, width=15, height=2, font=("仿宋GB_2312", 12)
        ).pack(pady=20)
        tk.Button(
            button_frame, text="文本输入",
            command=self.show_text_input_page, width=15, height=2, font=("仿宋GB_2312", 12)
        ).pack(pady=20)
        tk.Button(
            button_frame, text="生成模板表格",
            command=self.generate_empty_table, width=15, height=2, font=("仿宋GB_2312", 12)
        ).pack(pady=20)


def run_gui():
    root = tk.Tk()
    gui = LeaveApplicationGUI(root)
    gui.show_main_page()
    root.mainloop()


if __name__ == '__main__':
    run_gui()
